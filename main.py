from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import select, text
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import func

from io import BytesIO
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
import openpyxl
import re
import unicodedata

from .db import SessionLocal, engine, Base
from .models import Transportista, Flete
from .schemas import (
    TransportistaCreate,
    TransportistaOut,
    FleteCreate,
    FleteOut,
)

app = FastAPI(title="Logística Conecar API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],  # React (Vite)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Crea tablas (simple por ahora)
Base.metadata.create_all(bind=engine)

# Asegura que exista la columna "estado" en fletes (sin romper tu DB actual)
def ensure_estado_column():
    with engine.begin() as conn:
        conn.execute(text("""
            ALTER TABLE fletes
            ADD COLUMN IF NOT EXISTS estado VARCHAR(60);
        """))
        conn.execute(text("""
            CREATE INDEX IF NOT EXISTS ix_fletes_estado ON fletes (estado);
        """))

ensure_estado_column()


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@app.get("/health")
def health():
    return {"ok": True}

@app.get("/analytics")
def analytics(db: Session = Depends(get_db)):
    # Agrupado por mes
    by_mes = db.execute(
        select(
            Flete.anio_mes,
            func.coalesce(func.sum(Flete.flete_cobrado), 0).label("cobrado"),
            func.coalesce(func.sum(Flete.flete_pagado), 0).label("pagado"),
            func.coalesce(func.sum(Flete.diferencia), 0).label("diferencia"),
        )
        .where(Flete.anio_mes.isnot(None))
        .group_by(Flete.anio_mes)
        .order_by(Flete.anio_mes.asc())
    ).all()

    # Agrupado por estado
    by_estado = db.execute(
        select(
            Flete.estado,
            func.count(Flete.id).label("cantidad"),
            func.coalesce(func.sum(Flete.flete_cobrado), 0).label("cobrado"),
            func.coalesce(func.sum(Flete.flete_pagado), 0).label("pagado"),
            func.coalesce(func.sum(Flete.diferencia), 0).label("diferencia"),
        )
        .group_by(Flete.estado)
        .order_by(Flete.estado.asc())
    ).all()

    # Totales
    tot = db.execute(
        select(
            func.count(Flete.id).label("cantidad"),
            func.coalesce(func.sum(Flete.flete_cobrado), 0).label("cobrado"),
            func.coalesce(func.sum(Flete.flete_pagado), 0).label("pagado"),
            func.coalesce(func.sum(Flete.diferencia), 0).label("diferencia"),
        )
    ).one()

    def row_to_dict(r, keys):
        return {k: (float(getattr(r, k)) if k in ["cobrado","pagado","diferencia"] else getattr(r, k)) for k in keys}

    return {
        "totales": row_to_dict(tot, ["cantidad", "cobrado", "pagado", "diferencia"]),
        "por_mes": [row_to_dict(r, ["anio_mes", "cobrado", "pagado", "diferencia"]) for r in by_mes],
        "por_estado": [row_to_dict(r, ["estado", "cantidad", "cobrado", "pagado", "diferencia"]) for r in by_estado],
    }

# -------------------------
# Transportistas
# -------------------------
@app.post("/transportistas", response_model=TransportistaOut)
def crear_transportista(payload: TransportistaCreate, db: Session = Depends(get_db)):
    nombre = payload.nombre.strip()

    exists = db.execute(
        select(Transportista).where(Transportista.nombre == nombre)
    ).scalar_one_or_none()
    if exists:
        raise HTTPException(status_code=409, detail="Transportista ya existe")

    t = Transportista(nombre=nombre)
    db.add(t)
    db.commit()
    db.refresh(t)
    return t


@app.get("/transportistas", response_model=list[TransportistaOut])
def listar_transportistas(db: Session = Depends(get_db)):
    rows = db.execute(
        select(Transportista).order_by(Transportista.nombre.asc())
    ).scalars().all()
    return rows


# -------------------------
# Fletes
# -------------------------
@app.post("/fletes", response_model=FleteOut)
def crear_flete(payload: FleteCreate, db: Session = Depends(get_db)):
    o_carga = payload.o_carga.strip()

    exists = db.execute(
        select(Flete).where(Flete.o_carga == o_carga)
    ).scalar_one_or_none()
    if exists:
        raise HTTPException(status_code=409, detail="O.Carga ya existe")

    t = db.execute(
        select(Transportista).where(Transportista.id == payload.transportista_id)
    ).scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="Transportista no existe")

    cobrado = payload.flete_cobrado or 0
    pagado = payload.flete_pagado or 0
    diferencia = cobrado - pagado

    f = Flete(
        fecha=payload.fecha,
        dia=payload.dia,
        o_carga=o_carga,
        anio_mes=payload.anio_mes,
        cliente_destino=payload.cliente_destino,
        # estado NO se setea acá (se setea por import o más adelante por UI)
        transportista_id=payload.transportista_id,
        cod_transporte=payload.cod_transporte,
        ingrese_transporte=payload.ingrese_transporte,
        km=payload.km,
        tn_orden_carga=payload.tn_orden_carga,
        tn_cargadas=payload.tn_cargadas,
        aforo=payload.aforo,
        tarifa_asign=payload.tarifa_asign,
        flete_cobrado=payload.flete_cobrado,
        tarifa_tte=payload.tarifa_tte,
        flete_pagado=payload.flete_pagado,
        diferencia=diferencia,
        observacion=payload.observacion,
    )
    db.add(f)
    db.commit()
    db.refresh(f)
    return f


from sqlalchemy import or_

@app.get("/fletes", response_model=list[FleteOut])
def listar_fletes(
    estado: str | None = None,
    anio_mes: str | None = None,
    transportista_id: int | None = None,
    q: str | None = None,
    limit: int = 200,
    offset: int = 0,
    db: Session = Depends(get_db),
):
    limit = max(1, min(limit, 2000))
    offset = max(0, offset)

    stmt = select(Flete)

    if estado:
        stmt = stmt.where(Flete.estado == estado.strip().lower())

    if anio_mes:
        stmt = stmt.where(Flete.anio_mes == anio_mes.strip())

    if transportista_id:
        stmt = stmt.where(Flete.transportista_id == transportista_id)

    if q:
        qq = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                Flete.o_carga.ilike(qq),
                Flete.cliente_destino.ilike(qq),
            )
        )

    stmt = stmt.order_by(Flete.fecha.desc().nullslast(), Flete.id.desc()).limit(limit).offset(offset)

    rows = db.execute(stmt).scalars().all()
    return rows

from pydantic import BaseModel, Field

class FleteWebCreate(BaseModel):
    estado: str = Field(min_length=1, max_length=60)

    fecha: date | None = None
    dia: str | None = None
    o_carga: str
    anio_mes: str | None = None
    cliente_destino: str | None = None

    transportista_id: int

    cod_transporte: str | None = None
    ingrese_transporte: str | None = None

    km: Decimal | None = None
    tn_orden_carga: Decimal | None = None
    tn_cargadas: Decimal | None = None
    aforo: Decimal | None = None

    tarifa_asign: Decimal | None = None
    flete_cobrado: Decimal | None = None
    tarifa_tte: Decimal | None = None
    flete_pagado: Decimal | None = None

    observacion: str | None = None


@app.post("/fletes-web", response_model=FleteOut)
def crear_flete_web(payload: FleteWebCreate, db: Session = Depends(get_db)):
    oc = payload.o_carga.strip()
    if not oc:
        raise HTTPException(status_code=400, detail="O.Carga es obligatorio")

    exists = db.execute(select(Flete).where(Flete.o_carga == oc)).scalar_one_or_none()
    if exists:
        raise HTTPException(status_code=409, detail="O.Carga ya existe")

    validos = {"transporte", "viajes en camino", "viajes concretados"}
    est = payload.estado.strip().lower()
    if est not in validos:
        raise HTTPException(status_code=400, detail=f"Estado inválido. Usá: {sorted(validos)}")

    t = db.execute(select(Transportista).where(Transportista.id == payload.transportista_id)).scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="Transportista no existe")

    cobrado = payload.flete_cobrado or Decimal("0")
    pagado = payload.flete_pagado or Decimal("0")
    diferencia = cobrado - pagado

    f = Flete(
        estado=est,
        fecha=payload.fecha,
        dia=payload.dia,
        o_carga=oc,
        anio_mes=payload.anio_mes,
        cliente_destino=payload.cliente_destino,
        transportista_id=payload.transportista_id,
        cod_transporte=payload.cod_transporte,
        ingrese_transporte=payload.ingrese_transporte,
        km=payload.km,
        tn_orden_carga=payload.tn_orden_carga,
        tn_cargadas=payload.tn_cargadas,
        aforo=payload.aforo,
        tarifa_asign=payload.tarifa_asign,
        flete_cobrado=payload.flete_cobrado,
        tarifa_tte=payload.tarifa_tte,
        flete_pagado=payload.flete_pagado,
        diferencia=diferencia,
        observacion=payload.observacion,
    )
    db.add(f)
    db.commit()
    db.refresh(f)
    return f

# -------------------------
# Helpers import Excel
# -------------------------
def _norm(s: str) -> str:
    """
    Normaliza textos para matching:
    - quita tildes
    - MAYUS
    - reemplaza símbolos por espacios
    - colapsa espacios
    """
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.upper()
    s = re.sub(r"[./()\-\n\r\t]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


EXPECTED_COLS = {
    "FECHA": "fecha",
    "DIA": "dia",
    "O CARGA": "o_carga",
    "ANO MES": "anio_mes",
    "AÑO MES": "anio_mes",

    "CLIENTE DESTINO": "cliente_destino",
    "CLIENTE / DESTINO": "cliente_destino",
    "TRANSPORTISTA": "transportista",
    "COD TRANSPORTE": "cod_transporte",
    "INGRESE TRANSPORTE": "ingrese_transporte",

    "KM": "km",
    "TN ORDEN DE CARGA": "tn_orden_carga",
    "TN ORDEN CARGA": "tn_orden_carga",
    "TN CARGADAS": "tn_cargadas",
    "AFORO": "aforo",

    "TARIFA ASIGN": "tarifa_asign",
    "FLETE COBRADO": "flete_cobrado",
    "TARIFA TTE": "tarifa_tte",
    "TARIFA TTE.": "tarifa_tte",
    "FLETE PAGADO": "flete_pagado",

    "DIFERENCIA": "diferencia",
    "OBSERVACION": "observacion",
    "OBSERVACION ": "observacion",
    "OBSERVACIÓN": "observacion",
}


def _to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()

    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _to_decimal(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return None

    s = str(v).strip().replace(" ", "")
    # soporta 1.234,56 / 1234,56 / 1234.56
    if s.count(",") == 1 and s.count(".") >= 1:
        s = s.replace(".", "").replace(",", ".")
    elif s.count(",") == 1 and s.count(".") == 0:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _find_header_row_and_map(ws):
    """
    Busca una fila con headers que contengan al menos O.CARGA y TRANSPORTISTA.
    Devuelve: (header_row_idx, col_map) donde col_map = {field: col_index_1based}
    """
    for r in range(1, min(ws.max_row, 120) + 1):
        col_map = {}
        for c in range(1, ws.max_column + 1):
            raw = ws.cell(row=r, column=c).value
            if raw is None:
                continue
            key = _norm(raw)

            if key in EXPECTED_COLS:
                col_map[EXPECTED_COLS[key]] = c
                continue

            # fallback por contains
            if "CLIENTE" in key and "DESTINO" in key:
                col_map["cliente_destino"] = c

        if "o_carga" in col_map and "transportista" in col_map:
            return r, col_map

    return None, None


# -------------------------
# Import Excel (3 hojas) - SKIP por O.Carga
# -------------------------
@app.post("/import-excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)

    # Procesamos estas 3 hojas (si existen)
    sheets_to_process = {"transporte", "viajes en camino", "viajes concretados", "base datos"}

    inserted = 0
    skipped = 0
    transportistas_created = 0
    processed_sheets = []

    # Cache: O.Carga existentes
    existing = set(db.execute(select(Flete.o_carga)).scalars().all())

    # Cache transportistas por nombre
    transportista_cache = {}

    def get_or_create_transportista_id(nombre: str) -> int:
        nonlocal transportistas_created
        nombre = (nombre or "").strip()
        if not nombre:
            raise HTTPException(status_code=400, detail="Fila sin TRANSPORTISTA")

        if nombre in transportista_cache:
            return transportista_cache[nombre]

        t = db.execute(select(Transportista).where(Transportista.nombre == nombre)).scalar_one_or_none()
        if not t:
            t = Transportista(nombre=nombre)
            db.add(t)
            db.commit()
            db.refresh(t)
            transportistas_created += 1

        transportista_cache[nombre] = t.id
        return t.id

    def get_text(ws, r: int, col_map, field: str):
        if field not in col_map:
            return None
        v = ws.cell(row=r, column=col_map[field]).value
        if v is None:
            return None
        s = str(v).strip()
        return s if s != "" else None

    # Recorremos todas las hojas y tomamos solo las 3 que queremos
    for sname in wb.sheetnames:
        sname_clean = sname.strip().lower()
        if sname_clean not in sheets_to_process:
            continue

        ws = wb[sname]
        header_row, col_map = _find_header_row_and_map(ws)
        if not header_row:
            continue

        processed_sheets.append(sname)

        estado = "viajes concretados" if sname_clean == "base datos" else sname_clean

        r = header_row + 1
        blank_streak = 0
        MAX_BLANK_STREAK = 200  # corta cuando hay muchas filas vacías seguidas

        while True:
            oc_cell = ws.cell(row=r, column=col_map["o_carga"]).value
            o_carga = (str(oc_cell).strip() if oc_cell is not None else "")

            if not o_carga:
                blank_streak += 1
                if blank_streak >= MAX_BLANK_STREAK:
                    break
                r += 1
                continue

            blank_streak = 0

            # SKIP si ya existe
            if o_carga in existing:
                skipped += 1
                r += 1
                continue

            # Transportista
            tr_name = ws.cell(row=r, column=col_map["transportista"]).value
            transportista_id = get_or_create_transportista_id(str(tr_name) if tr_name is not None else "")

            fecha = _to_date(ws.cell(row=r, column=col_map["fecha"]).value) if "fecha" in col_map else None

            km = _to_decimal(ws.cell(row=r, column=col_map["km"]).value) if "km" in col_map else None
            tn_orden = _to_decimal(ws.cell(row=r, column=col_map["tn_orden_carga"]).value) if "tn_orden_carga" in col_map else None
            tn_carg = _to_decimal(ws.cell(row=r, column=col_map["tn_cargadas"]).value) if "tn_cargadas" in col_map else None
            aforo = _to_decimal(ws.cell(row=r, column=col_map["aforo"]).value) if "aforo" in col_map else None

            tarifa_asign = _to_decimal(ws.cell(row=r, column=col_map["tarifa_asign"]).value) if "tarifa_asign" in col_map else None
            flete_cobrado = _to_decimal(ws.cell(row=r, column=col_map["flete_cobrado"]).value) if "flete_cobrado" in col_map else None
            tarifa_tte = _to_decimal(ws.cell(row=r, column=col_map["tarifa_tte"]).value) if "tarifa_tte" in col_map else None
            flete_pagado = _to_decimal(ws.cell(row=r, column=col_map["flete_pagado"]).value) if "flete_pagado" in col_map else None

            cobrado = flete_cobrado or Decimal("0")
            pagado = flete_pagado or Decimal("0")
            diferencia = cobrado - pagado

            f = Flete(
                fecha=fecha,
                dia=get_text(ws, r, col_map, "dia"),
                o_carga=o_carga,
                anio_mes=get_text(ws, r, col_map, "anio_mes"),
                cliente_destino=get_text(ws, r, col_map, "cliente_destino"),
                estado=estado,
                transportista_id=transportista_id,
                cod_transporte=get_text(ws, r, col_map, "cod_transporte"),
                ingrese_transporte=get_text(ws, r, col_map, "ingrese_transporte"),
                km=km,
                tn_orden_carga=tn_orden,
                tn_cargadas=tn_carg,
                aforo=aforo,
                tarifa_asign=tarifa_asign,
                flete_cobrado=flete_cobrado,
                tarifa_tte=tarifa_tte,
                flete_pagado=flete_pagado,
                diferencia=diferencia,
                observacion=get_text(ws, r, col_map, "observacion"),
            )

            db.add(f)
            inserted += 1
            existing.add(o_carga)

            if inserted % 200 == 0:
                db.commit()

            r += 1

    db.commit()

    if not processed_sheets:
        raise HTTPException(status_code=400, detail="No encontré ninguna de las 3 hojas objetivo para importar.")

    return {
        "ok": True,
        "processed_sheets": processed_sheets,
        "inserted": inserted,
        "skipped": skipped,
        "transportistas_created": transportistas_created,
    }


# -------------------------
# Export Excel (3 hojas)
# -------------------------
@app.get("/export-excel")
def export_excel(db: Session = Depends(get_db)):
    headers = [
        "FECHA",
        "Día",
        "O.Carga",
        "AÑO.MES",
        "CLIENTE / DESTINO",
        "TRANSPORTISTA",
        "Cod. Transporte",
        "INGRESE TRANSPORTE",
        "KM",
        "TN ORDEN DE CARGA",
        "TN CARGADAS",
        "AFORO",
        "TARIFA ASIGN",
        "FLETE COBRADO",
        "TARIFA TTE.",
        "FLETE PAGADO",
        "DIFERENCIA",
        "OBSERVACION",
    ]

    wb = openpyxl.Workbook()
    # Sacamos la hoja default
    wb.remove(wb.active)

    # Cache transportistas
    transportistas = db.execute(select(Transportista)).scalars().all()
    tmap = {t.id: t.nombre for t in transportistas}

    def dec_to_number(x):
        if x is None:
            return None
        try:
            return float(x)
        except Exception:
            return x

    def add_sheet(title: str, estado_value: str):
        ws = wb.create_sheet(title=title)
        ws.append(headers)

        fletes = db.execute(
            select(Flete)
            .where(Flete.estado == estado_value)
            .order_by(Flete.fecha.asc().nullslast(), Flete.o_carga.asc())
        ).scalars().all()

        for f in fletes:
            transportista_nombre = tmap.get(f.transportista_id, "")

            cobrado = f.flete_cobrado or Decimal("0")
            pagado = f.flete_pagado or Decimal("0")
            diferencia = cobrado - pagado

            ws.append([
                f.fecha,
                f.dia,
                f.o_carga,
                f.anio_mes,
                f.cliente_destino,
                transportista_nombre,
                f.cod_transporte,
                f.ingrese_transporte,
                dec_to_number(f.km),
                dec_to_number(f.tn_orden_carga),
                dec_to_number(f.tn_cargadas),
                dec_to_number(f.aforo),
                dec_to_number(f.tarifa_asign),
                dec_to_number(f.flete_cobrado),
                dec_to_number(f.tarifa_tte),
                dec_to_number(f.flete_pagado),
                dec_to_number(diferencia),
                f.observacion,
            ])

    add_sheet("transporte", "transporte")
    add_sheet("viajes en camino", "viajes en camino")
    add_sheet("viajes concretados", "viajes concretados")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "FLETES_COBRADOS_PAGADOS_EXPORT.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
from pydantic import BaseModel, Field

class EstadoUpdate(BaseModel):
    estado: str = Field(min_length=1, max_length=60)

@app.patch("/fletes/{o_carga}/estado")
def cambiar_estado(o_carga: str, payload: EstadoUpdate, db: Session = Depends(get_db)):
    oc = o_carga.strip()

    f = db.execute(select(Flete).where(Flete.o_carga == oc)).scalar_one_or_none()
    if not f:
        raise HTTPException(status_code=404, detail="No existe ese O.Carga")

    nuevo = payload.estado.strip().lower()
    validos = {"transporte", "viajes en camino", "viajes concretados"}
    if nuevo not in validos:
        raise HTTPException(status_code=400, detail=f"Estado inválido. Usá: {sorted(validos)}")

    f.estado = nuevo
    db.commit()

    return {"ok": True, "o_carga": oc, "estado": nuevo}
